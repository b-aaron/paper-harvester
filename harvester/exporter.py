"""
exporter.py - Saves scraped articles to Excel and organises PDF files.

The Excel workbook contains one sheet per journal (or one combined sheet)
with all available metadata fields as columns.

PDF files are renamed to 'FirstAuthorLastName Year JournalAbbr Title.pdf'
and saved to the user-specified directory.
"""

import os
import re
import shutil
from typing import Dict, List, Optional

from .scrapers.base_scraper import Article


class Exporter:
    """
    Exports Article objects to an Excel workbook and renames/organises PDFs.

    Usage::

        exporter = Exporter(save_dir="/path/to/output")
        exporter.to_excel(articles, filename="papers.xlsx")
        exporter.organise_pdfs(articles)
    """

    # Excel column headers in display order
    COLUMNS = [
        "title",
        "authors",
        "journal",
        "year",
        "month",
        "volume",
        "issue",
        "pages",
        "doi",
        "url",
        "abstract",
        "keywords",
        "publisher",
        "pdf_path",
        "oa_pdf_url",
    ]

    def __init__(self, save_dir: str = "."):
        """
        Initialize the exporter.

        Args:
            save_dir: Root directory where output files are written.
        """
        self.save_dir = save_dir
        os.makedirs(save_dir, exist_ok=True)

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def to_excel(
        self,
        articles: List[Article],
        filename: str = "papers.xlsx",
        group_by_journal: bool = True,
    ) -> str:
        """
        Write article metadata to an Excel workbook.

        Args:
            articles: List of Article objects to export.
            filename: Output filename (placed inside save_dir).
            group_by_journal: If True, each journal gets its own worksheet.
                              If False, all articles go into a single sheet.

        Returns:
            Absolute path to the created Excel file.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            ) from exc

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove the default empty sheet

        if group_by_journal:
            # Group articles by journal
            groups: Dict[str, List[Article]] = {}
            for art in articles:
                groups.setdefault(art.journal or "Unknown", []).append(art)
            for journal_name, arts in groups.items():
                sheet_name = self._safe_sheet_name(journal_name)
                ws = wb.create_sheet(title=sheet_name)
                self._write_sheet(ws, arts)
        else:
            ws = wb.create_sheet(title="Articles")
            self._write_sheet(ws, articles)

        path = os.path.join(self.save_dir, filename)
        wb.save(path)
        print(f"  ✓ Excel saved: {path}")
        return path

    def _write_sheet(self, ws, articles: List[Article]) -> None:
        """
        Write headers and article rows to a worksheet, with basic formatting.

        Args:
            ws: openpyxl Worksheet object.
            articles: Articles to write.
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        HEADER_FONT = Font(bold=True, color="FFFFFF")

        # Header row
        display_names = {
            "title": "Title",
            "authors": "Authors",
            "journal": "Journal",
            "year": "Year",
            "month": "Month",
            "volume": "Volume",
            "issue": "Issue",
            "pages": "Pages",
            "doi": "DOI",
            "url": "URL",
            "abstract": "Abstract",
            "keywords": "Keywords",
            "publisher": "Publisher",
            "pdf_path": "PDF Path",
            "oa_pdf_url": "OA PDF URL",
        }
        for col_idx, col_key in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=display_names.get(col_key, col_key))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row_idx, article in enumerate(articles, 2):
            row_data = article.to_dict()
            for col_idx, col_key in enumerate(self.COLUMNS, 1):
                value = row_data.get(col_key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=(col_key == "abstract"))

        # Auto-adjust column widths (capped at 80 chars for readability)
        for col_idx, col_key in enumerate(self.COLUMNS, 1):
            col_letter = get_column_letter(col_idx)
            # Abstract and title get a fixed wider column
            if col_key in ("abstract", "title"):
                ws.column_dimensions[col_letter].width = 60
            elif col_key == "authors":
                ws.column_dimensions[col_letter].width = 35
            else:
                ws.column_dimensions[col_letter].width = 20

        # Freeze top row
        ws.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # PDF organisation
    # ------------------------------------------------------------------

    def organise_pdfs(self, articles: List[Article], pdf_dir: Optional[str] = None) -> None:
        """
        Rename and move all downloaded PDFs to a structured directory.

        Target filename format: '{FirstAuthor} {Year} {JournalAbbr} {Title}.pdf'

        Args:
            articles: Articles whose ``pdf_path`` should be organised.
            pdf_dir: Destination directory for PDFs. Defaults to save_dir/pdfs.
        """
        dest_dir = pdf_dir or os.path.join(self.save_dir, "pdfs")
        os.makedirs(dest_dir, exist_ok=True)

        for article in articles:
            if not article.pdf_path or not os.path.isfile(article.pdf_path):
                continue
            new_name = self._build_pdf_name(article)
            dest_path = os.path.join(dest_dir, new_name)
            try:
                shutil.move(article.pdf_path, dest_path)
                article.pdf_path = dest_path
                print(f"  ✓ PDF moved: {new_name}")
            except OSError as exc:
                print(f"  [WARN] Could not move PDF: {exc}")

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _build_pdf_name(article: Article) -> str:
        """
        Build a descriptive, filesystem-safe PDF filename.

        Format: '{LastName} {Year} {Journal} {Title}.pdf'

        Args:
            article: Source article.

        Returns:
            Filename string ending in '.pdf'.
        """
        first_author = (
            article.authors[0].split()[-1] if article.authors else "Unknown"
        )
        year = str(article.year) if article.year else "0000"
        journal = (article.journal or "Journal")[:25]
        title = (article.title or "NoTitle")[:60]

        name = f"{first_author} {year} {journal} {title}"
        # Remove characters illegal in most filesystems
        name = re.sub(r'[\\/:*?"<>|]', " ", name)
        # Collapse whitespace
        name = re.sub(r"\s+", " ", name).strip()
        return name + ".pdf"

    @staticmethod
    def _safe_sheet_name(name: str) -> str:
        """
        Return an Excel-safe worksheet name (max 31 chars, no special chars).

        Args:
            name: Raw sheet name candidate.

        Returns:
            Sanitised sheet name string.
        """
        safe = re.sub(r"[\\/:*?\[\]]", "_", name)
        return safe[:31]
